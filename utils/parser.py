"""
parser.py
Lê o arquivo .xlsx do cronograma de projetos e retorna:
  - df: DataFrame normalizado (Projeto, Atividade, Responsável, Semana, Horas)
  - ferias: dict {pessoa: [lista de datas (semanas) de férias]}
  - semanas: lista ordenada de datas das semanas
  - pessoas: lista de nomes únicos
  - projetos: lista de projetos únicos
"""

import re
import pandas as pd
from openpyxl import load_workbook


def _is_green(cell):
    """Retorna True se a célula tem fundo verde (férias)."""
    fill = cell.fill
    if fill is None:
        return False
    color = fill.fgColor
    if color is None:
        return False
    # Cor em RGB hex
    if color.type == "rgb":
        rgb = color.rgb  # ex: "FF92D050" — ARGB
        # Remove alpha channel
        hex_color = rgb[-6:].upper()
        if len(hex_color) == 6:
            r = int(hex_color[0:2], 16)
            g = int(hex_color[2:4], 16)
            b = int(hex_color[4:6], 16)
            # Verde: G dominante, R e B menores
            if g > 150 and g > r + 30 and g > b + 30:
                return True
    return False


def _parse_date_header(value):
    """Tenta interpretar um cabeçalho de semana como data."""
    if value is None:
        return None
    # Já é datetime/date — caso mais comum quando vem do Excel via openpyxl
    import datetime as _dt
    if isinstance(value, (_dt.datetime, _dt.date)):
        return pd.Timestamp(value)
    # Fallback: string como "6/4", "06/04", "6/4/2025"
    s = str(value).strip()
    match = re.match(r"^(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?$", s)
    if match:
        d, m, y = match.groups()
        year = int(y) if y else 2026
        if year < 100:
            year += 2000
        try:
            return pd.Timestamp(year=year, month=int(m), day=int(d))
        except Exception:
            return None
    return None


def parse_cronograma(file_bytes: bytes):
    """
    Parâmetros:
        file_bytes: conteúdo binário do arquivo .xlsx

    Retorna:
        df, ferias, semanas, pessoas, projetos
    """
    import io
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active  # considera a primeira aba

    rows = list(ws.iter_rows())
    if not rows:
        raise ValueError("Planilha vazia.")

    # -------------------------------------------------------
    # 1. Identificar linha de cabeçalho (que contém "Responsável")
    # -------------------------------------------------------
    header_row_idx = None
    col_projeto = None
    col_atividade = None
    col_responsavel = None
    semana_cols = []  # lista de (col_idx, date)

    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            val = str(cell.value).strip().lower() if cell.value else ""
            if "responsável" in val or "responsavel" in val:
                header_row_idx = i
                col_responsavel = j
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        raise ValueError("Coluna 'Responsável' não encontrada na planilha.")

    header_row = rows[header_row_idx]

    # Identificar colunas Projeto e Atividade (estão antes de Responsável)
    for j in range(col_responsavel):
        val = str(header_row[j].value).strip().lower() if header_row[j].value else ""
        if "projeto" in val:
            col_projeto = j
        elif "atividade" in val or "tarefa" in val:
            col_atividade = j

    # Se não encontrou por nome, assume posição 0=Projeto, 1=Atividade
    if col_projeto is None:
        col_projeto = 0
    if col_atividade is None:
        col_atividade = 1

    # Colunas de semanas (após Responsável)
    for j in range(col_responsavel + 1, len(header_row)):
        cell = header_row[j]
        dt = _parse_date_header(cell.value)
        if dt is not None:
            semana_cols.append((j, dt))

    if not semana_cols:
        raise ValueError("Nenhuma coluna de semana (datas) encontrada na planilha.")

    # -------------------------------------------------------
    # 2. Iterar linhas de dados
    # -------------------------------------------------------
    records = []
    ferias = {}  # {pessoa: set de timestamps}
    projeto_atual = None

    data_rows = rows[header_row_idx + 1:]

    for row in data_rows:
        # Projeto: herda se célula vazia
        proj_val = row[col_projeto].value
        if proj_val and str(proj_val).strip():
            projeto_atual = str(proj_val).strip()

        if projeto_atual is None:
            continue

        atividade = row[col_atividade].value
        if not atividade or not str(atividade).strip():
            continue
        atividade = str(atividade).strip()

        responsavel = row[col_responsavel].value
        if not responsavel or not str(responsavel).strip():
            continue
        responsavel = str(responsavel).strip()

        for (col_idx, semana_dt) in semana_cols:
            if col_idx >= len(row):
                continue
            cell = row[col_idx]

            # Detectar célula verde = férias
            if _is_green(cell):
                if responsavel not in ferias:
                    ferias[responsavel] = set()
                ferias[responsavel].add(semana_dt)

            # Extrair horas
            valor = cell.value
            horas = 0.0
            if valor is not None:
                try:
                    horas = float(valor)
                except (ValueError, TypeError):
                    pass

            if horas > 0:
                records.append({
                    "Projeto": projeto_atual,
                    "Atividade": atividade,
                    "Responsável": responsavel,
                    "Semana": semana_dt,
                    "Horas": horas,
                })

    if not records:
        raise ValueError("Nenhum dado de horas encontrado na planilha.")

    df = pd.DataFrame(records)
    df["Semana"] = pd.to_datetime(df["Semana"])
    df = df.sort_values(["Projeto", "Semana", "Responsável"]).reset_index(drop=True)

    semanas = sorted(df["Semana"].unique().tolist())
    pessoas = sorted(df["Responsável"].unique().tolist())
    projetos = sorted(df["Projeto"].unique().tolist())

    # Converter sets para listas ordenadas
    ferias = {p: sorted(list(v)) for p, v in ferias.items()}

    return df, ferias, semanas, pessoas, projetos
